"""
EnterpriseIntegrationHub - Enterprise Connector for Google/Microsoft APIs

This module connects to Google Workspace and Microsoft 365 APIs.
Fetches document histories, revision logs, and file contents for comprehensive analysis.
"""

import io
import re
import json
import logging
from typing import Dict, Any, Optional, List
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class EnterpriseIntegrationHub:
    """
    Connects to Google Workspace and Microsoft 365 APIs.
    Fetches document histories, revision logs, and file contents.
    """
    
    def __init__(self):
        self.google_creds = None
        self.ms_creds = None
        self.google_drive = None
        self.google_docs = None
        self.google_sheets = None
        self.microsoft_graph = None
        
        # API connection status
        self.connection_status = {
            "google": False,
            "microsoft": False
        }
    
    def connect_google(self, credentials: Dict[str, Any]) -> bool:
        """
        Connect to Google Workspace APIs.
        
        Args:
            credentials: Dictionary containing Google OAuth credentials
                        (token, refresh_token, client_id, client_secret, etc.)
            
        Returns:
            True if connection successful, False otherwise
        """
        try:
            # Try to import Google API libraries
            try:
                from google.oauth2.credentials import Credentials
                from googleapiclient.discovery import build
                
                # Create credentials object
                self.google_creds = Credentials.from_authorized_user_info(credentials)
                
                # Build API clients
                self.google_drive = build('drive', 'v3', credentials=self.google_creds)
                self.google_docs = build('docs', 'v1', credentials=self.google_creds)
                self.google_sheets = build('sheets', 'v4', credentials=self.google_creds)
                
                self.connection_status["google"] = True
                logger.info("Successfully connected to Google Workspace APIs")
                return True
                
            except ImportError as e:
                logger.error(f"Google API libraries not available: {str(e)}")
                self.connection_status["google"] = False
                return False
                
        except Exception as e:
            logger.error(f"Failed to connect to Google Workspace: {str(e)}")
            self.connection_status["google"] = False
            return False
    
    def connect_microsoft(self, tenant_id: str, client_id: str, client_secret: str) -> bool:
        """
        Connect to Microsoft 365 APIs.
        
        Args:
            tenant_id: Microsoft 365 tenant ID
            client_id: Application client ID
            client_secret: Application client secret
            
        Returns:
            True if connection successful, False otherwise
        """
        try:
            # Try to import Microsoft Graph libraries
            try:
                from msal import ConfidentialClientApplication
                import requests
                
                # Create MSAL application
                app = ConfidentialClientApplication(
                    client_id,
                    authority=f"https://login.microsoftonline.com/{tenant_id}",
                    client_credential=client_secret
                )
                
                # Get access token
                result = app.acquire_token_for_client(
                    scopes=["https://graph.microsoft.com/.default"]
                )
                
                if 'access_token' in result:
                    # Create requests session with authorization header
                    self.microsoft_graph = requests.Session()
                    self.microsoft_graph.headers.update({
                        'Authorization': f"Bearer {result['access_token']}",
                        'Content-Type': 'application/json'
                    })
                    
                    self.connection_status["microsoft"] = True
                    logger.info("Successfully connected to Microsoft Graph API")
                    return True
                else:
                    logger.error(f"Failed to acquire Microsoft token: {result.get('error_description')}")
                    self.connection_status["microsoft"] = False
                    return False
                    
            except ImportError as e:
                logger.error(f"Microsoft Graph libraries not available: {str(e)}")
                self.connection_status["microsoft"] = False
                return False
                
        except Exception as e:
            logger.error(f"Failed to connect to Microsoft 365: {str(e)}")
            self.connection_status["microsoft"] = False
            return False
    
    def fetch_google_doc_with_revisions(self, doc_id: str) -> Dict[str, Any]:
        """
        Fetches a Google Doc and its entire revision history.
        
        Args:
            doc_id: Google Doc ID
            
        Returns:
            Dictionary containing document content and revision history
        """
        if not self.connection_status["google"]:
            return {
                "error": "Google Workspace not connected",
                "content": "",
                "revision_history": []
            }
        
        try:
            # Get the document content
            doc = self.google_docs.documents().get(documentId=doc_id).execute()
            content = self._extract_text_from_google_doc(doc)
            
            # Get revision history from Drive API
            try:
                revisions = self.google_drive.revisions().list(
                    fileId=doc_id,
                    fields='revisions(id,modifiedTime,lastModifyingUser,size)'
                ).execute()
                
                revision_history = []
                for rev in revisions.get('revisions', []):
                    revision_history.append({
                        'timestamp': rev.get('modifiedTime'),
                        'author': rev.get('lastModifyingUser', {}).get('displayName', 'Unknown'),
                        'version': rev.get('id'),
                        'size': rev.get('size', 0)
                    })
            except Exception as e:
                logger.warning(f"Could not fetch revision history: {str(e)}")
                revision_history = []
            
            return {
                'content': content,
                'revision_history': revision_history,
                'full_analysis': self._analyze_revision_timeline(revision_history),
                'doc_metadata': {
                    'title': doc.get('title', 'Untitled'),
                    'document_id': doc_id
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching Google Doc: {str(e)}")
            return {
                "error": str(e),
                "content": "",
                "revision_history": []
            }
    
    def fetch_google_sheet_with_revisions(self, sheet_id: str) -> Dict[str, Any]:
        """
        Fetches a Google Sheet and its revision history.
        
        Args:
            sheet_id: Google Sheet ID
            
        Returns:
            Dictionary containing sheet data and revision history
        """
        if not self.connection_status["google"]:
            return {
                "error": "Google Workspace not connected",
                "data": {},
                "revision_history": []
            }
        
        try:
            # Get spreadsheet metadata
            spreadsheet = self.google_sheets.spreadsheets().get(
                spreadsheetId=sheet_id
            ).execute()
            
            # Get all sheets data
            data = {}
            for sheet in spreadsheet.get('sheets', []):
                sheet_title = sheet['properties']['title']
                sheet_id = sheet['properties']['sheetId']
                
                # Get sheet values
                result = self.google_sheets.spreadsheets().values().get(
                    spreadsheetId=sheet_id,
                    range=sheet_title
                ).execute()
                
                data[sheet_title] = result.get('values', [])
            
            # Get revision history
            try:
                revisions = self.google_drive.revisions().list(
                    fileId=sheet_id,
                    fields='revisions(id,modifiedTime,lastModifyingUser)'
                ).execute()
                
                revision_history = []
                for rev in revisions.get('revisions', []):
                    revision_history.append({
                        'timestamp': rev.get('modifiedTime'),
                        'author': rev.get('lastModifyingUser', {}).get('displayName', 'Unknown'),
                        'version': rev.get('id')
                    })
            except Exception as e:
                logger.warning(f"Could not fetch revision history: {str(e)}")
                revision_history = []
            
            return {
                'data': data,
                'revision_history': revision_history,
                'spreadsheet_metadata': {
                    'title': spreadsheet.get('properties', {}).get('title', 'Untitled'),
                    'sheet_count': len(data)
                }
            }
            
        except Exception as e:
            logger.error(f"Error fetching Google Sheet: {str(e)}")
            return {
                "error": str(e),
                "data": {},
                "revision_history": []
            }
    
    def fetch_microsoft_word_with_revisions(self, file_id: str) -> Dict[str, Any]:
        """
        Fetches a Microsoft Word document and its revision history.
        
        Args:
            file_id: Microsoft Graph file ID
            
        Returns:
            Dictionary containing document content and version history
        """
        if not self.connection_status["microsoft"]:
            return {
                "error": "Microsoft 365 not connected",
                "content": "",
                "versions": []
            }
        
        try:
            # Download the file
            response = self.microsoft_graph.get(
                f"https://graph.microsoft.com/v1.0/drives/me/items/{file_id}/content"
            )
            
            if response.status_code == 200:
                file_bytes = response.content
                
                # Get version history
                versions_response = self.microsoft_graph.get(
                    f"https://graph.microsoft.com/v1.0/drives/me/items/{file_id}/versions"
                )
                
                versions = versions_response.json().get('value', []) if versions_response.status_code == 200 else []
                
                # Extract text from Word document (simplified)
                content = self._extract_text_from_word_bytes(file_bytes)
                
                return {
                    'content': content,
                    'raw_content': file_bytes,
                    'versions': versions,
                    'analysis': self._analyze_word_revisions(versions)
                }
            else:
                return {
                    "error": f"Failed to download file: {response.status_code}",
                    "content": "",
                    "versions": []
                }
                
        except Exception as e:
            logger.error(f"Error fetching Microsoft Word document: {str(e)}")
            return {
                "error": str(e),
                "content": "",
                "versions": []
            }
    
    def fetch_microsoft_excel_with_versions(self, file_id: str) -> Dict[str, Any]:
        """
        Fetches Microsoft Excel file with version history.
        
        Args:
            file_id: Microsoft Graph file ID
            
        Returns:
            Dictionary containing Excel data and version history
        """
        if not self.connection_status["microsoft"]:
            return {
                "error": "Microsoft 365 not connected",
                "data": {},
                "versions": []
            }
        
        try:
            # Download the file
            response = self.microsoft_graph.get(
                f"https://graph.microsoft.com/v1.0/drives/me/items/{file_id}/content"
            )
            
            if response.status_code == 200:
                file_bytes = response.content
                
                # Get version history
                versions_response = self.microsoft_graph.get(
                    f"https://graph.microsoft.com/v1.0/drives/me/items/{file_id}/versions"
                )
                
                versions = versions_response.json().get('value', []) if versions_response.status_code == 200 else []
                
                # Parse Excel data (simplified CSV extraction)
                data = self._extract_data_from_excel_bytes(file_bytes)
                
                return {
                    'data': data,
                    'raw_content': file_bytes,
                    'versions': versions,
                    'analysis': self._analyze_excel_versions(versions)
                }
            else:
                return {
                    "error": f"Failed to download file: {response.status_code}",
                    "data": {},
                    "versions": []
                }
                
        except Exception as e:
            logger.error(f"Error fetching Microsoft Excel file: {str(e)}")
            return {
                "error": str(e),
                "data": {},
                "versions": []
            }
    
    def _extract_text_from_google_doc(self, doc: Dict) -> str:
        """Extracts plain text from Google Doc JSON structure"""
        try:
            content = doc.get('body', {}).get('content', [])
            text_parts = []
            
            for element in content:
                if 'paragraph' in element:
                    para = element['paragraph']
                    for elem in para.get('elements', []):
                        if 'textRun' in elem:
                            text_parts.append(elem['textRun'].get('content', ''))
                elif 'table' in element:
                    # Handle tables
                    table = element['table']
                    for row in table.get('tableRows', []):
                        row_text = []
                        for cell in row.get('tableCells', []):
                            cell_text = self._extract_text_from_google_doc({'body': {'content': cell.get('content', [])}})
                            row_text.append(cell_text.strip())
                        text_parts.append(' | '.join(row_text))
            
            return ''.join(text_parts)
        except Exception as e:
            logger.error(f"Error extracting text from Google Doc: {str(e)}")
            return ""
    
    def _extract_text_from_word_bytes(self, file_bytes: bytes) -> str:
        """Extract text from Word document bytes"""
        try:
            # Try to use python-docx if available
            try:
                from docx import Document
                import io
                
                doc_stream = io.BytesIO(file_bytes)
                doc = Document(doc_stream)
                
                text_parts = []
                for paragraph in doc.paragraphs:
                    text_parts.append(paragraph.text)
                
                return '\n'.join(text_parts)
                
            except ImportError:
                # Fallback: try to extract readable text
                text = file_bytes.decode('utf-8', errors='ignore')
                # Remove binary artifacts
                text = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\xff]', '', text)
                return text
                
        except Exception as e:
            logger.error(f"Error extracting text from Word: {str(e)}")
            return ""
    
    def _extract_data_from_excel_bytes(self, file_bytes: bytes) -> Dict[str, List]:
        """Extract data from Excel bytes"""
        try:
            # Try to use openpyxl if available
            try:
                from openpyxl import load_workbook
                import io
                
                wb_stream = io.BytesIO(file_bytes)
                wb = load_workbook(wb_stream, data_only=True)
                
                data = {}
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    data[sheet_name] = []
                    for row in sheet.iter_rows(values_only=True):
                        data[sheet_name].append(row)
                
                return data
                
            except ImportError:
                # Fallback: return empty structure
                return {"error": "Excel parsing library not available"}
                
        except Exception as e:
            logger.error(f"Error extracting data from Excel: {str(e)}")
            return {"error": str(e)}
    
    def _analyze_revision_timeline(self, revisions: List[Dict]) -> Dict[str, Any]:
        """Analyzes revision timeline for paste events and typing patterns"""
        if len(revisions) < 2:
            return {
                "paste_detected": False,
                "editing_pattern": "unknown",
                "total_revisions": len(revisions)
            }
        
        try:
            # Calculate time differences between revisions
            time_intervals = []
            size_changes = []
            
            for i in range(1, len(revisions)):
                prev_rev = revisions[i-1]
                curr_rev = revisions[i]
                
                # Parse timestamps
                try:
                    prev_time = datetime.fromisoformat(prev_rev['timestamp'].replace('Z', '+00:00'))
                    curr_time = datetime.fromisoformat(curr_rev['timestamp'].replace('Z', '+00:00'))
                    time_diff = (curr_time - prev_time).total_seconds()
                    time_intervals.append(time_diff)
                except:
                    pass
                
                # Track size changes
                prev_size = prev_rev.get('size', 0)
                curr_size = curr_rev.get('size', 0)
                size_changes.append(curr_size - prev_size)
            
            # Detect paste events (large sudden additions)
            paste_detected = any(change > 1000 for change in size_changes)
            
            # Determine editing pattern
            if time_intervals:
                avg_interval = sum(time_intervals) / len(time_intervals)
                if avg_interval < 60:
                    editing_pattern = "rapid"
                elif avg_interval < 300:
                    editing_pattern = "continuous"
                else:
                    editing_pattern = "sporadic"
            else:
                editing_pattern = "unknown"
            
            return {
                "paste_detected": paste_detected,
                "editing_pattern": editing_pattern,
                "total_revisions": len(revisions),
                "avg_time_interval": avg_interval if time_intervals else 0,
                "size_changes": size_changes
            }
            
        except Exception as e:
            logger.error(f"Error analyzing revision timeline: {str(e)}")
            return {
                "paste_detected": False,
                "editing_pattern": "error",
                "error": str(e)
            }
    
    def _analyze_word_revisions(self, versions: List[Dict]) -> Dict[str, Any]:
        """Analyzes Word document version history"""
        return {
            "total_versions": len(versions),
            "version_analysis": "Word version analysis not fully implemented"
        }
    
    def _analyze_excel_versions(self, versions: List[Dict]) -> Dict[str, Any]:
        """Analyzes Excel version history"""
        return {
            "total_versions": len(versions),
            "version_analysis": "Excel version analysis not fully implemented"
        }
    
    def get_connection_status(self) -> Dict[str, bool]:
        """Get current connection status for all platforms"""
        return self.connection_status.copy()


class MockIntegrationHub(EnterpriseIntegrationHub):
    """
    Mock implementation for testing without actual API credentials.
    Returns simulated data for development and testing.
    """
    
    def fetch_google_doc_with_revisions(self, doc_id: str) -> Dict[str, Any]:
        """Mock implementation returning simulated data"""
        return {
            'content': "This is a sample document content for testing purposes. " +
                      "It contains multiple sentences to simulate a real document. " +
                      "The analysis should detect if this appears to be AI-generated.",
            'revision_history': [
                {
                    'timestamp': '2024-01-15T10:00:00Z',
                    'author': 'student@example.com',
                    'version': '1'
                },
                {
                    'timestamp': '2024-01-15T10:30:00Z',
                    'author': 'student@example.com',
                    'version': '2'
                }
            ],
            'full_analysis': {
                'paste_detected': False,
                'editing_pattern': 'continuous',
                'total_revisions': 2
            },
            'doc_metadata': {
                'title': 'Sample Document',
                'document_id': doc_id
            }
        }
    
    def fetch_google_sheet_with_revisions(self, sheet_id: str) -> Dict[str, Any]:
        """Mock implementation returning simulated data"""
        return {
            'data': {
                'Sheet1': [
                    ['Name', 'Score', 'Grade'],
                    ['Alice', '95', 'A'],
                    ['Bob', '87', 'B']
                ]
            },
            'revision_history': [
                {
                    'timestamp': '2024-01-15T10:00:00Z',
                    'author': 'student@example.com',
                    'version': '1'
                }
            ],
            'spreadsheet_metadata': {
                'title': 'Sample Sheet',
                'sheet_count': 1
            }
        }
    
    def fetch_microsoft_word_with_revisions(self, file_id: str) -> Dict[str, Any]:
        """Mock implementation returning simulated data"""
        return {
            'content': "This is a sample Word document content for testing.",
            'raw_content': b'Mock Word content',
            'versions': [
                {
                    'id': '1',
                    'lastModifiedDateTime': '2024-01-15T10:00:00Z'
                }
            ],
            'analysis': {
                'total_versions': 1,
                'version_analysis': 'Mock analysis'
            }
        }
    
    def fetch_microsoft_excel_with_versions(self, file_id: str) -> Dict[str, Any]:
        """Mock implementation returning simulated data"""
        return {
            'data': {
                'Sheet1': [
                    ['Data1', 'Data2'],
                    ['Value1', 'Value2']
                ]
            },
            'raw_content': b'Mock Excel content',
            'versions': [
                {
                    'id': '1',
                    'lastModifiedDateTime': '2024-01-15T10:00:00Z'
                }
            ],
            'analysis': {
                'total_versions': 1,
                'version_analysis': 'Mock analysis'
            }
        }
